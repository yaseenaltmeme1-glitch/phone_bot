# bot.py — دليل أرقام المستشفى (بالعربي) + بصمة إنكليزية + احصائيات احترافية (Admin فقط)
import os, logging, asyncio, math, re, sqlite3
import io, csv
from typing import Dict, List, Tuple, Optional
from datetime import datetime, timedelta

try:
    from zoneinfo import ZoneInfo
    BAGHDAD_TZ = ZoneInfo("Asia/Baghdad")
except Exception:
    BAGHDAD_TZ = None

from openpyxl import load_workbook
from telegram import (
    Update,
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    InputFile
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler,
    ContextTypes, filters
)
from telegram.error import RetryAfter

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.getenv("DATA_DIR", BASE)

# ==================== بصمتك ====================
SIGNATURE  = "\n────────────\nSource: CCTV – Yaseen Al-Tamimi"

# نص "عن البوت"
ABOUT_TEXT = (
    "ℹ️ عن البوت\n"
    "بوت دليل أرقام المستشفى، يوفّر بحث ذكي ويعرض النتائج بشكل مبسط وسريع.\n\n"
    "📩 لمزيد من الاستفسارات أو مقترحات التعديل:\n"
    "@ya_se91\n\n"
    "────────────\n"
    "Source: CCTV – Yaseen Al-Tamimi"
)

# ============= Admin Stats Settings =============
ADMIN_ID = 8099482759  # 👑 فقط هذا الـID يطلع احصائيات
DB_PATH = os.path.join(BASE, "stats.db")

# أسماء أعمدة محتملة
DEPT_CANDIDATES  = ["القسم","قسم","الاسم","اسم القسم"]
PHONE_CANDIDATES = ["رقم الهاتف","الهاتف","رقم","موبايل","Phone"]

# ذاكرة
display_rows: List[Tuple[str, str]] = []
departments:  List[str] = []
phonebook:    Dict[str, str] = {}

# كيبورد رئيسية
MAIN_KB = ReplyKeyboardMarkup(
    [
        [KeyboardButton("📞 أرقام المستشفى")],
        [KeyboardButton("🔍 بحث بالاسم")],
        [KeyboardButton("ℹ️ عن البوت")],
        [KeyboardButton("◀️ رجوع للقائمة")]
    ],
    resize_keyboard=True
)

# إعداد الشبكات
GRID_COLS      = 3
PAGE_SIZE_ALL  = 24
PAGE_SIZE_SRCH = 21

# ---------------- تطبيع ----------------
ARABIC_DIAC = re.compile(r"[ًٌٍَُِّْـ]")

def strip_diacritics(s: str) -> str:
    return ARABIC_DIAC.sub("", s or "")

def normalize_arabic(s: str) -> str:
    s = str(s or "")
    s = s.replace("\u200f","").replace("\u200e","").replace("\ufeff","").strip()
    s = strip_diacritics(s)
    s = s.replace("آ","ا").replace("أ","ا").replace("إ","ا")
    s = s.replace("ى","ي").replace("ة","ه")
    s = re.sub(r"[^\w\s\u0600-\u06FF]"," ", s)
    s = re.sub(r"\s+"," ", s).strip()
    return s.upper()

# ---------------- وقت بغداد ----------------
def now_baghdad() -> datetime:
    if BAGHDAD_TZ:
        return datetime.now(BAGHDAD_TZ)
    return datetime.utcnow() + timedelta(hours=3)

def iso(dt: datetime) -> str:
    return dt.replace(microsecond=0).isoformat()

def period_bounds(kind: str) -> Tuple[Optional[datetime], Optional[datetime], str]:
    """returns (start, end, title). if start/end None => all-time"""
    now = now_baghdad()
    if kind == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, now, "📊 إحصائيات اليوم"
    if kind == "week":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=now.weekday())
        return start, now, "📅 إحصائيات هذا الأسبوع"
    if kind == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now, "🗓️ إحصائيات هذا الشهر"
    if kind == "7":
        return now - timedelta(days=7), now, "📆 آخر 7 أيام"
    if kind == "30":
        return now - timedelta(days=30), now, "📆 آخر 30 يوم"
    if kind == "90":
        return now - timedelta(days=90), now, "📆 آخر 90 يوم"
    return None, None, "♾️ إحصائيات من البداية"

# ---------------- DB (SQLite) ----------------
def db_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn

def init_db():
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            first_seen TEXT NOT NULL,
            last_seen  TEXT NOT NULL,
            username   TEXT,
            full_name  TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            chat_id INTEGER,
            event_type TEXT NOT NULL,
            dept TEXT,
            query TEXT,
            extra TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_ts ON events(ts)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_type ON events(event_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_dept ON events(dept)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_user ON events(user_id)")
    conn.commit()
    conn.close()

def upsert_user(user) -> None:
    if not user:
        return
    uid = user.id
    username = user.username or ""
    full_name = (user.full_name or "").strip()
    t = iso(now_baghdad())
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT user_id FROM users WHERE user_id=?", (uid,))
    row = cur.fetchone()
    if row:
        cur.execute(
            "UPDATE users SET last_seen=?, username=?, full_name=? WHERE user_id=?",
            (t, username, full_name, uid)
        )
    else:
        cur.execute(
            "INSERT INTO users(user_id, first_seen, last_seen, username, full_name) VALUES(?,?,?,?,?)",
            (uid, t, t, username, full_name)
        )
    conn.commit()
    conn.close()

def log_event(event_type: str, user_id: int, chat_id: Optional[int], dept: str = "", query: str = "", extra: str = "") -> None:
    t = iso(now_baghdad())
    conn = db_conn()
    conn.execute(
        "INSERT INTO events(ts, user_id, chat_id, event_type, dept, query, extra) VALUES(?,?,?,?,?,?,?)",
        (t, user_id, chat_id if chat_id is not None else None, event_type, dept or "", query or "", extra or "")
    )
    conn.commit()
    conn.close()

def is_admin(update: Update) -> bool:
    u = update.effective_user
    return bool(u and u.id == ADMIN_ID)

# ---------------- قراءة الإكسل ----------------
def list_excel_files(folder: str) -> List[str]:
    try:
        return [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    except:
        return []

def read_headers(ws) -> List[str]:
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c or "").strip() for c in row]
    return []

def find_col_idx(headers: List[str], candidates: List[str]) -> Optional[int]:
    H = [normalize_arabic(h) for h in headers]
    C = [normalize_arabic(c) for c in candidates]
    for i,h in enumerate(H):
        if h in C: return i
    for i,h in enumerate(H):
        for c in C:
            if c in h: return i
    return None

def load_phonebook() -> Tuple[int,str]:
    global display_rows, departments, phonebook
    display_rows, departments, phonebook = [], [], {}
    files = list_excel_files(DATA_DIR)
    if not files:
        return 0, f"❌ ماكو ملفات ‎.xlsx داخل: {DATA_DIR}"
    total = 0
    for path in files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = read_headers(ws)
            if not headers: 
                wb.close(); 
                continue
            di = find_col_idx(headers, DEPT_CANDIDATES)
            pi = find_col_idx(headers, PHONE_CANDIDATES)
            if di is None or pi is None: 
                wb.close(); 
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: 
                    continue
                dept  = str(row[di] if di < len(row) and row[di] is not None else "").strip()
                phone = str(row[pi] if pi < len(row) and row[pi] is not None else "").strip()
                if not dept: 
                    continue
                display_rows.append((dept, phone))
                phonebook[normalize_arabic(dept)] = phone
                total += 1
            wb.close()
        except Exception as e:
            logging.exception(f"Load error in {path}: {e}")
    display_rows.sort(key=lambda x: x[0])
    departments = [d for d,_ in display_rows]
    return total, (f"✅ تم تحميل {total} سجل." if total else "❌ لم يتم تحميل أي سجل.")

# ---------------- أدوات إرسال ----------------
async def safe_reply(update: Update, text: str, reply_markup=None):
    text = f"{text}{SIGNATURE}"
    try:
        return await update.message.reply_text(text, reply_markup=reply_markup)
    except RetryAfter as e:
        await asyncio.sleep(e.retry_after + 1)
        return await update.message.reply_text(text, reply_markup=reply_markup)

async def reply_plain(msg, text: str, reply_markup=None):
    text = f"{text}{SIGNATURE}"
    try:
        return await msg.reply_text(text, reply_markup=reply_markup)
    except RetryAfter as e:
        await asyncio.sleep(e.retry_after + 1)
        return await msg.reply_text(text, reply_markup=reply_markup)

async def safe_edit(q, text: str, reply_markup=None):
    try:
        return await q.message.edit_text(text, reply_markup=reply_markup)
    except RetryAfter as e:
        await asyncio.sleep(e.retry_after + 1)
        return await q.message.edit_text(text, reply_markup=reply_markup)

# ---------------- الانترو ----------------
def build_intro() -> str:
    return (
        "👋 أهلاً بك في بوت أرقام المستشفى.\n\n"
        "📌 طريقة الاستخدام:\n"
        "• 📞 أرقام المستشفى: تصفّح الأقسام كمربعات.\n"
        "• 🔍 بحث بالاسم: اكتب أي جزء من اسم القسم.\n"
        "• ℹ️ عن البوت: معلومات عن البوت.\n"
        "• ◀️ رجوع: العودة إلى هذه القائمة.\n\n"
        "────────────\n"
        "Source: CCTV – Yaseen Al-Tamimi"
    )

# ---------------- الشبكات ----------------
def build_grid(indices: List[int], page: int, page_size: int, cols: int, mode: str) -> InlineKeyboardMarkup:
    total = len(indices)
    pages = max(1, math.ceil(total / page_size))
    page  = max(0, min(page, pages-1))
    start, end = page*page_size, min(page*page_size + page_size, total)
    slice_idx = indices[start:end]

    rows, row = [], []
    for idx in slice_idx:
        name = departments[idx]
        row.append(InlineKeyboardButton(name, callback_data=f"dept:{idx}"))
        if len(row) == cols:
            rows.append(row); row = []
    if row: rows.append(row)

    if pages > 1:
        ctrl = []
        if page > 0:             ctrl.append(InlineKeyboardButton("⬅️ السابق", callback_data=f"{mode}:{page-1}"))
        ctrl.append(InlineKeyboardButton(f"صفحة {page+1}/{pages}", callback_data="noop"))
        if page < pages-1:       ctrl.append(InlineKeyboardButton("التالي ➡️", callback_data=f"{mode}:{page+1}"))
        rows.append(ctrl)
    rows.append([InlineKeyboardButton("◀️ رجوع للقائمة", callback_data="home")])
    return InlineKeyboardMarkup(rows)

def grid_all(page:int=0) -> InlineKeyboardMarkup:
    return build_grid(list(range(len(departments))), page, PAGE_SIZE_ALL, GRID_COLS, "allp")

def grid_search(matches: List[int], page:int=0) -> InlineKeyboardMarkup:
    return build_grid(matches, page, PAGE_SIZE_SRCH, GRID_COLS, "srchp")

# ---------------- البحث ----------------
def search_indices(query: str) -> List[int]:
    qn = normalize_arabic(query)
    if not qn: return []
    matches = []
    for i, name in enumerate(departments):
        if qn in normalize_arabic(name):
            matches.append(i)
    return matches


# ---------------- Admin لوحة الاحصائيات ----------------
DISPLAY_CITY = os.getenv("DISPLAY_CITY", "Karbala")

def admin_menu() -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("📊 اليوم",  callback_data="stats:today"),
         InlineKeyboardButton("📅 أسبوع", callback_data="stats:week"),
         InlineKeyboardButton("🗓️ شهر",  callback_data="stats:month")],
        [InlineKeyboardButton("♾️ من البداية", callback_data="stats:all")],

        [InlineKeyboardButton("🏆 Top 10 أقسام (من البداية)", callback_data="adm:top10_depts")],

        [InlineKeyboardButton("👥 عدد المستخدمين الكلي", callback_data="adm:user_count"),
         InlineKeyboardButton("👥 Top 15 مستخدم", callback_data="adm:top15_users_all")],

        [InlineKeyboardButton("🧾 آخر 25 مستخدم (مع الوقت)", callback_data="adm:recent25")],
        [InlineKeyboardButton("🕒 آخر نشاط", callback_data="adm:last_activity")],

        [InlineKeyboardButton("📥 تصدير التقارير", callback_data="export:menu")],
        [InlineKeyboardButton("📣 إرسال رسالة ترحيب للمستخدمين", callback_data="broadcast:menu")],

        [InlineKeyboardButton("◀️ رجوع للقائمة", callback_data="home")]
    ]
    return InlineKeyboardMarkup(rows)

def export_menu() -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("📄 تقرير اليوم (XLSX)", callback_data="export:xlsx:today"),
         InlineKeyboardButton("📄 تقرير اليوم (CSV)", callback_data="export:csv:today")],

        [InlineKeyboardButton("📄 تقرير كلي (XLSX)", callback_data="export:xlsx:all"),
         InlineKeyboardButton("📄 تقرير كلي (CSV)", callback_data="export:csv:all")],

        [InlineKeyboardButton("📄 بحث + مستخدمين (XLSX)", callback_data="export:xlsx:usage"),
         InlineKeyboardButton("📄 بحث + مستخدمين (CSV)", callback_data="export:csv:usage")],

        [InlineKeyboardButton("📄 تقرير شامل (XLSX)", callback_data="export:xlsx:full"),
         InlineKeyboardButton("📄 تقرير شامل (CSV)", callback_data="export:csv:full")],

        [InlineKeyboardButton("◀️ رجوع للإحصائيات", callback_data="admin:home")]
    ]
    return InlineKeyboardMarkup(rows)

def broadcast_menu() -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("✅ إرسال الآن", callback_data="broadcast:confirm"),
         InlineKeyboardButton("❌ إلغاء", callback_data="broadcast:cancel")],
        [InlineKeyboardButton("◀️ رجوع للإحصائيات", callback_data="admin:home")]
    ]
    return InlineKeyboardMarkup(rows)

def _where_ts(start: datetime, end: Optional[datetime]) -> Tuple[str, Tuple]:
    return "WHERE ts >= ? AND ts <= ?", (iso(start), iso(end))

def fmt_ts(ts: str) -> str:
    if not ts:
        return "—"
    try:
        dt = datetime.fromisoformat(ts)
    except Exception:
        return ts
    try:
        if BAGHDAD_TZ:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=BAGHDAD_TZ)
            else:
                dt = dt.astimezone(BAGHDAD_TZ)
    except Exception:
        pass
    return dt.strftime("%Y-%m-%d  %H:%M:%S")

def get_total_users() -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    total = cur.fetchone()[0] or 0
    conn.close()
    return total

def get_last_activity_ts() -> str:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT MAX(ts) FROM events")
    ts = cur.fetchone()[0] or ""
    conn.close()
    return fmt_ts(ts) if ts else "—"

def period_bounds(kind: str) -> Tuple[datetime, datetime, str]:
    now = now_baghdad()
    if kind == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, now, "📊 إحصائيات اليوم"
    if kind == "week":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=now.weekday())
        return start, now, "📅 إحصائيات هذا الأسبوع"
    if kind == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now, "🗓️ إحصائيات هذا الشهر"
    # fallback
    start = now - timedelta(days=7)
    return start, now, "📆 آخر 7 أيام"

def stats_summary(kind: str) -> str:
    total_users = get_total_users()
    last_activity = get_last_activity_ts()

    if kind == "all":
        conn = db_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(DISTINCT user_id) FROM events")
        active_users = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM events WHERE event_type IN ('search_text','dept_select','search_hit')")
        total_search = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM events WHERE event_type='search_text'")
        total_text_search = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM events WHERE event_type='dept_select'")
        total_button_search = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM events WHERE event_type='search_hit'")
        total_hit_search = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM events WHERE event_type='not_found'")
        not_found = cur.fetchone()[0] or 0

        conn.close()

        return (
            "♾️ إحصائيات من البداية\n"
            f"• 👥 مجموع المستخدمين الكلي: {total_users}\n"
            f"• ✅ مستخدمين نشطين (من البداية): {active_users}\n"
            f"• 🔎 مجموع عمليات البحث: {total_search}\n"
            f"   - ✍️ بحث كتابة (محاولات): {total_text_search}\n"
            f"   - 🧩 اختيار قسم (أزرار): {total_button_search}\n"
            f"   - ✅ تطابق مباشر (بحث واحد): {total_hit_search}\n"
            f"• ❌ بدون نتيجة: {not_found}\n"
            f"• 🕒 آخر نشاط ({DISPLAY_CITY}): {last_activity}"
        )

    start, end, title = period_bounds(kind)
    where, params = _where_ts(start, end)

    conn = db_conn()
    cur = conn.cursor()

    cur.execute(
        "SELECT COUNT(*) FROM users WHERE first_seen >= ? AND first_seen <= ?",
        (iso(start), iso(end))
    )
    new_users = cur.fetchone()[0] or 0

    cur.execute(
        f"SELECT COUNT(DISTINCT user_id) FROM events {where}",
        params
    )
    active_users = cur.fetchone()[0] or 0

    cur.execute(
        f"SELECT COUNT(*) FROM events {where} AND event_type IN ('search_text','dept_select','search_hit')",
        params
    )
    total_search = cur.fetchone()[0] or 0

    cur.execute(
        f"SELECT COUNT(*) FROM events {where} AND event_type='search_text'",
        params
    )
    total_text_search = cur.fetchone()[0] or 0

    cur.execute(
        f"SELECT COUNT(*) FROM events {where} AND event_type='dept_select'",
        params
    )
    total_button_search = cur.fetchone()[0] or 0

    cur.execute(
        f"SELECT COUNT(*) FROM events {where} AND event_type='search_hit'",
        params
    )
    total_hit_search = cur.fetchone()[0] or 0

    cur.execute(
        f"SELECT COUNT(*) FROM events {where} AND event_type='not_found'",
        params
    )
    not_found = cur.fetchone()[0] or 0

    conn.close()

    return (
        f"{title}\n"
        f"• 👥 مجموع المستخدمين الكلي: {total_users}\n"
        f"• 👤 مستخدمين جدد: {new_users}\n"
        f"• ✅ مستخدمين نشطين: {active_users}\n"
        f"• 🔎 عمليات البحث: {total_search}\n"
        f"   - ✍️ بحث كتابة (محاولات): {total_text_search}\n"
        f"   - 🧩 اختيار قسم (أزرار): {total_button_search}\n"
        f"   - ✅ تطابق مباشر (بحث واحد): {total_hit_search}\n"
        f"• ❌ بدون نتيجة: {not_found}\n"
        f"• 🕒 آخر نشاط ({DISPLAY_CITY}): {last_activity}"
    )

def top10_departments_alltime() -> str:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT dept, COUNT(*) AS c
        FROM events
        WHERE event_type IN ('dept_select','search_hit') AND dept <> ''
        GROUP BY dept
        ORDER BY c DESC
        LIMIT 10
    """)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return "🏆 Top 10 أقسام (من البداية)\n❌ لا توجد بيانات كافية بعد."
    lines = ["🏆 Top 10 أقسام (من البداية)"]
    for i, (dept, c) in enumerate(rows, 1):
        lines.append(f"{i}) {dept} — {c}")
    return "\n".join(lines)

def top15_users_alltime_detailed() -> str:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, COUNT(*) AS c
        FROM events
        WHERE event_type IN ('dept_select','search_hit','search_text')
        GROUP BY user_id
        ORDER BY c DESC
        LIMIT 15
    """)
    rows = cur.fetchall()

    result = []
    for uid, c in rows:
        cur.execute("SELECT full_name, username, last_seen FROM users WHERE user_id=?", (uid,))
        urow = cur.fetchone()
        full_name = (urow[0] if urow and urow[0] else "").strip()
        username = (urow[1] if urow and urow[1] else "").strip()
        last_seen = fmt_ts(urow[2]) if urow and urow[2] else "—"
        label = full_name if full_name else (f"@{username}" if username else str(uid))
        handle = f"@{username}" if username else "—"
        result.append((uid, label, handle, last_seen, c))

    conn.close()

    if not result:
        return "👥 Top 15 مستخدم (من البداية)\n❌ لا توجد بيانات كافية بعد."

    lines = ["👥 Top 15 مستخدم (من البداية)"]
    for i, (uid, name, handle, last_seen, c) in enumerate(result, 1):
        lines.append(f"{i}) {name} | {handle} | آخر استخدام: {last_seen} | الاستخدام: {c}")
    return "\n".join(lines)

def recent25_users() -> str:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, full_name, username, last_seen, first_seen
        FROM users
        ORDER BY last_seen DESC
        LIMIT 25
    """)
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return "🧾 آخر 25 مستخدم\n❌ لا توجد بيانات بعد."

    lines = ["🧾 آخر 25 مستخدم (آخر استخدام)"]
    for i, (uid, full_name, username, last_seen, first_seen) in enumerate(rows, 1):
        name = (full_name or "").strip() or str(uid)
        handle = f"@{username}" if username else "—"
        last_s = fmt_ts(last_seen) if last_seen else "—"
        first_s = fmt_ts(first_seen) if first_seen else "—"
        lines.append(f"{i}) {name} | {handle} | آخر: {last_s} | أول: {first_s}")
    return "\n".join(lines)

def export_report_bytes(fmt: str, scope: str) -> Tuple[bytes, str, str]:
    """
    fmt: 'csv' or 'xlsx'
    scope: today|all|usage|full
    returns (bytes, filename, mime)
    """
    if fmt == "csv":
        data = build_csv(scope)
        filename = f"report_{scope}.csv"
        mime = "text/csv"
        return data, filename, mime

    data = build_xlsx(scope)
    filename = f"report_{scope}.xlsx"
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return data, filename, mime

def build_csv(scope: str) -> bytes:
    import io, csv
    out = io.StringIO()
    w = csv.writer(out)

    total_users = get_total_users()
    last_activity = get_last_activity_ts()
    gen = fmt_ts(iso(now_baghdad()))

    def section(title):
        w.writerow([])
        w.writerow([title])

    # SUMMARY
    w.writerow(["GeneratedAt", f"{gen} ({DISPLAY_CITY})"])
    w.writerow(["TotalUsers", total_users])
    w.writerow(["LastActivity", f"{last_activity} ({DISPLAY_CITY})"])

    if scope in ("today", "full"):
        start, end, _ = period_bounds("today")
        section("TODAY SUMMARY")
        w.writerow(["From", fmt_ts(iso(start))])
        w.writerow(["To", fmt_ts(iso(end))])
        w.writerow(["NewUsersToday", count_new_users(start, end)])
        w.writerow(["ActiveUsersToday", count_active_users(start, end)])
        w.writerow(["SearchesToday", count_searches(start, end)])
        w.writerow(["NotFoundToday", count_not_found(start, end)])

    if scope in ("all", "full"):
        section("ALL-TIME SUMMARY")
        w.writerow(["ActiveUsersAllTime", count_active_users_all()])
        w.writerow(["SearchesAllTime", count_searches_all()])
        w.writerow(["NotFoundAllTime", count_not_found_all()])

    if scope in ("usage", "full"):
        section("TOP 10 DEPARTMENTS (ALL-TIME)")
        w.writerow(["Rank", "Department", "Count"])
        for i, (dept, c) in enumerate(get_top_depts(10), 1):
            w.writerow([i, dept, c])

        section("TOP 15 USERS (ALL-TIME)")
        w.writerow(["Rank", "UserID", "Name", "Username", "LastSeen", "Count"])
        for i, r in enumerate(get_top_users(15), 1):
            uid, name, username, last_seen, c = r
            w.writerow([i, uid, name, username, last_seen, c])

        section("RECENT 25 USERS (BY LAST SEEN)")
        w.writerow(["Rank", "UserID", "Name", "Username", "FirstSeen", "LastSeen"])
        for i, r in enumerate(get_recent_users(25), 1):
            uid, name, username, first_seen, last_seen = r
            w.writerow([i, uid, name, username, first_seen, last_seen])

    return out.getvalue().encode("utf-8-sig")

def build_xlsx(scope: str) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    def set_cols(sheet, widths):
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w

    header_font = Font(bold=True)
    center = Alignment(horizontal="center")

    total_users = get_total_users()
    last_activity = get_last_activity_ts()
    gen = fmt_ts(iso(now_baghdad()))

    # Summary sheet
    ws.append(["Report", "PhoneBook Bot"])
    ws.append(["GeneratedAt", f"{gen} ({DISPLAY_CITY})"])
    ws.append(["TotalUsers", total_users])
    ws.append(["LastActivity", f"{last_activity} ({DISPLAY_CITY})"])
    for r in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=2):
        for cell in r:
            cell.font = header_font if cell.column == 1 else cell.font
    set_cols(ws, [22, 55])

    if scope in ("today", "full"):
        ws.append([])
        ws.append(["Today From", fmt_ts(iso(period_bounds("today")[0]))])
        ws.append(["Today To", fmt_ts(iso(period_bounds("today")[1]))])
        start, end, _ = period_bounds("today")
        ws.append(["NewUsersToday", count_new_users(start, end)])
        ws.append(["ActiveUsersToday", count_active_users(start, end)])
        ws.append(["SearchesToday", count_searches(start, end)])
        ws.append(["NotFoundToday", count_not_found(start, end)])
        set_cols(ws, [22, 55])

    if scope in ("all", "full"):
        ws.append([])
        ws.append(["ActiveUsersAllTime", count_active_users_all()])
        ws.append(["SearchesAllTime", count_searches_all()])
        ws.append(["NotFoundAllTime", count_not_found_all()])

    # Usage sheets
    if scope in ("usage", "full"):
        ws2 = wb.create_sheet("TopDepts")
        ws2.append(["Rank", "Department", "Count"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.alignment = center
        for i, (dept, c) in enumerate(get_top_depts(10), 1):
            ws2.append([i, dept, c])
        set_cols(ws2, [8, 45, 10])

        ws3 = wb.create_sheet("TopUsers")
        ws3.append(["Rank", "UserID", "Name", "Username", "LastSeen", "Count"])
        for cell in ws3[1]:
            cell.font = header_font
            cell.alignment = center
        for i, (uid, name, username, last_seen, c) in enumerate(get_top_users(15), 1):
            ws3.append([i, uid, name, username, last_seen, c])
        set_cols(ws3, [8, 14, 28, 18, 22, 10])

        ws4 = wb.create_sheet("Recent25")
        ws4.append(["Rank", "UserID", "Name", "Username", "FirstSeen", "LastSeen"])
        for cell in ws4[1]:
            cell.font = header_font
            cell.alignment = center
        for i, (uid, name, username, first_seen, last_seen) in enumerate(get_recent_users(25), 1):
            ws4.append([i, uid, name, username, first_seen, last_seen])
        set_cols(ws4, [8, 14, 28, 18, 22, 22])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def get_top_depts(limit: int = 10) -> List[Tuple[str, int]]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT dept, COUNT(*) AS c
        FROM events
        WHERE event_type IN ('dept_select','search_hit') AND dept <> ''
        GROUP BY dept
        ORDER BY c DESC
        LIMIT ?
    """, (limit,))
    rows = [(d, int(c)) for d, c in cur.fetchall() if d]
    conn.close()
    return rows

def get_top_users(limit: int = 15) -> List[Tuple[int, str, str, str, int]]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, COUNT(*) AS c
        FROM events
        WHERE event_type IN ('dept_select','search_hit','search_text')
        GROUP BY user_id
        ORDER BY c DESC
        LIMIT ?
    """, (limit,))
    rows = cur.fetchall()

    result = []
    for uid, c in rows:
        cur.execute("SELECT full_name, username, last_seen FROM users WHERE user_id=?", (uid,))
        urow = cur.fetchone()
        full_name = (urow[0] if urow and urow[0] else "").strip()
        username = (urow[1] if urow and urow[1] else "").strip()
        last_seen = fmt_ts(urow[2]) if urow and urow[2] else "—"
        name = full_name if full_name else str(uid)
        handle = f"@{username}" if username else "—"
        result.append((uid, name, handle, last_seen, int(c)))

    conn.close()
    return result

def get_recent_users(limit: int = 25) -> List[Tuple[int, str, str, str, str]]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, full_name, username, first_seen, last_seen
        FROM users
        ORDER BY last_seen DESC
        LIMIT ?
    """, (limit,))
    rows = []
    for uid, full_name, username, first_seen, last_seen in cur.fetchall():
        name = (full_name or "").strip() or str(uid)
        handle = f"@{username}" if username else "—"
        rows.append((uid, name, handle, fmt_ts(first_seen), fmt_ts(last_seen)))
    conn.close()
    return rows

def count_new_users(start: datetime, end: datetime) -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users WHERE first_seen >= ? AND first_seen <= ?", (iso(start), iso(end)))
    n = cur.fetchone()[0] or 0
    conn.close()
    return int(n)

def count_active_users(start: datetime, end: datetime) -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(DISTINCT user_id) FROM events WHERE ts >= ? AND ts <= ?", (iso(start), iso(end)))
    n = cur.fetchone()[0] or 0
    conn.close()
    return int(n)

def count_searches(start: datetime, end: datetime) -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) FROM events 
        WHERE ts >= ? AND ts <= ? AND event_type IN ('search_text','dept_select','search_hit')
    """, (iso(start), iso(end)))
    n = cur.fetchone()[0] or 0
    conn.close()
    return int(n)

def count_not_found(start: datetime, end: datetime) -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM events WHERE ts >= ? AND ts <= ? AND event_type='not_found'", (iso(start), iso(end)))
    n = cur.fetchone()[0] or 0
    conn.close()
    return int(n)

def count_active_users_all() -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(DISTINCT user_id) FROM events")
    n = cur.fetchone()[0] or 0
    conn.close()
    return int(n)

def count_searches_all() -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM events WHERE event_type IN ('search_text','dept_select','search_hit')")
    n = cur.fetchone()[0] or 0
    conn.close()
    return int(n)

def count_not_found_all() -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM events WHERE event_type='not_found'")
    n = cur.fetchone()[0] or 0
    conn.close()
    return int(n)

def broadcast_template() -> str:
    return (
        "👋 تحية طيبة من مستشفى الإمام الحسن المجتبى (عليه السلام)\n\n"
        "نحب نسمع اقتراحاتكم أو أي تعديل تحبون نضيفه للبوت حتى يكون أدق وأسهل.\n"
        "اكتبوا لنا أفكاركم بكل راحة 🙏\n\n"
        "مع الشكر والتقدير 🌿"
    )
# ---------------- Handlers ----------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("start", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    await update.message.reply_text(build_intro(), reply_markup=MAIN_KB)

async def about_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("about", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    await safe_reply(update, ABOUT_TEXT, reply_markup=MAIN_KB)

async def reload_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("reload", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    n,msg = load_phonebook()
    await safe_reply(update, msg, reply_markup=MAIN_KB)

async def admin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    log_event("admin_open", update.effective_user.id, update.effective_chat.id if update.effective_chat else None)
    if not is_admin(update):
        await safe_reply(update, "⛔️ غير مصرح.", reply_markup=MAIN_KB)
        return
    await safe_reply(update, "👑 لوحة الإحصائيات (للأدمن فقط):", reply_markup=admin_menu())

async def list_depts(update: Update, context: ContextTypes.DEFAULT_TYPE, page:int=0):
    if not departments:
        await safe_reply(update, "❌ لا توجد سجلات. استخدم /reload بعد التأكد من ملف الإكسل.", reply_markup=MAIN_KB)
        return
    await reply_plain(update.message, "اختر القسم من القائمة:", reply_markup=grid_all(page))

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    upsert_user(update.effective_user)
    txt = (update.message.text or "").strip()
    chat_id = update.effective_chat.id if update.effective_chat else None
    uid = update.effective_user.id

    if txt == "📞 أرقام المستشفى":
        log_event("open_list", uid, chat_id)
        await list_depts(update, context, 0)
        return

    if txt == "🔍 بحث بالاسم":
        log_event("prompt_search", uid, chat_id)
        await safe_reply(update, "✍️ اكتب أي جزء من اسم القسم.", reply_markup=MAIN_KB)
        return

    if txt == "ℹ️ عن البوت":
        log_event("about_btn", uid, chat_id)
        await safe_reply(update, ABOUT_TEXT, reply_markup=MAIN_KB)
        return

    if txt == "◀️ رجوع للقائمة":
        log_event("back_home", uid, chat_id)
        await safe_reply(update, build_intro(), reply_markup=MAIN_KB)
        return

    matches = search_indices(txt)
    log_event("search_text", uid, chat_id, query=txt, extra=f"matches={len(matches)}")

    if not matches:
        log_event("not_found", uid, chat_id, query=txt)
        await safe_reply(update, "❌ لم يتم العثور على هذا القسم.", reply_markup=MAIN_KB)
        return

    if len(matches) == 1:
        idx = matches[0]
        name = departments[idx]
        num = phonebook.get(normalize_arabic(name), "")
        log_event("search_hit", uid, chat_id, dept=name, query=txt)
        await safe_reply(update, f"✅ {name} — {num if num else '—'}", reply_markup=MAIN_KB)
        return

    context.user_data["last_search_indices"] = matches
    await reply_plain(update.message, "🔎 تم العثور على عدة نتائج، اختر القسم:", reply_markup=grid_search(matches, 0))

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    data = q.data if q else ""
    uid = update.effective_user.id if update.effective_user else None
    chat_id = update.effective_chat.id if update.effective_chat else None

    try:        

        # ADMIN callbacks
        if data.startswith(("stats:", "adm:", "export:", "broadcast:", "admin:")):
            await q.answer()
            if not (update.effective_user and update.effective_user.id == ADMIN_ID):
                await reply_plain(q.message, "⛔️ غير مصرح.", reply_markup=MAIN_KB)
                return

            # رجوع للوحة الأدمن
            if data == "admin:home":
                await reply_plain(q.message, "👑 لوحة الإحصائيات (للأدمن فقط):", reply_markup=admin_menu())
                return

            # إحصائيات (تبقى مثل ما هي)
            if data.startswith("stats:"):
                kind = data.split(":", 1)[1]
                if kind not in ("today", "week", "month", "all"):
                    kind = "all"
                text = stats_summary(kind)
                await reply_plain(q.message, text, reply_markup=admin_menu())
                return

            # Top 10 أقسام (من البداية)
            if data == "adm:top10_depts":
                text = top10_departments_alltime()
                await reply_plain(q.message, text, reply_markup=admin_menu())
                return

            # عدد المستخدمين الكلي
            if data == "adm:user_count":
                total = get_total_users()
                last_act = get_last_activity_ts()
                text = (
                    "👥 عدد المستخدمين الكلي\n"
                    f"• المجموع: {total}\n"
                    f"• آخر نشاط ({DISPLAY_CITY}): {last_act}"
                )
                await reply_plain(q.message, text, reply_markup=admin_menu())
                return

            # Top 15 مستخدم (من البداية) + تفاصيل
            if data == "adm:top15_users_all":
                text = top15_users_alltime_detailed()
                await reply_plain(q.message, text, reply_markup=admin_menu())
                return

            # آخر 25 مستخدم
            if data == "adm:recent25":
                text = recent25_users()
                await reply_plain(q.message, text, reply_markup=admin_menu())
                return

            # آخر نشاط فقط
            if data == "adm:last_activity":
                last_act = get_last_activity_ts()
                text = f"🕒 آخر نشاط ({DISPLAY_CITY})\n{last_act}"
                await reply_plain(q.message, text, reply_markup=admin_menu())
                return

            # قائمة التصدير
            if data == "export:menu":
                await reply_plain(q.message, "📥 اختر نوع التقرير للتصدير:", reply_markup=export_menu())
                return

            # تصدير
            if data.startswith("export:"):
                # export:fmt:scope
                parts = data.split(":")
                if len(parts) == 3 and parts[1] in ("csv", "xlsx") and parts[2] in ("today", "all", "usage", "full"):
                    fmt, scope = parts[1], parts[2]
                    file_bytes, filename, mime = export_report_bytes(fmt, scope)

                    bio = io.BytesIO(file_bytes)
                    bio.name = filename
                    await q.message.reply_document(document=InputFile(bio, filename=filename), caption="✅ تم تصدير التقرير.")
                    return

                await reply_plain(q.message, "❌ خيار تصدير غير صالح.", reply_markup=export_menu())
                return

            # رسالة ترحيب للمستخدمين
            if data == "broadcast:menu":
                preview = broadcast_template()
                await q.message.reply_text(
                    "📣 سيتم إرسال الرسالة التالية إلى جميع مستخدمي البوت:\n\n"
                    f"{preview}\n\n"
                    "اختر:",
                    reply_markup=broadcast_menu()
                )
                return

            if data == "broadcast:cancel":
                await reply_plain(q.message, "✅ تم إلغاء الإرسال.", reply_markup=admin_menu())
                return

            if data == "broadcast:confirm":
                # إرسال فعلي (Admin فقط) مع احترام Flood control
                preview = broadcast_template()
                conn = db_conn()
                cur = conn.cursor()
                cur.execute("SELECT user_id FROM users")
                user_ids = [r[0] for r in cur.fetchall()]
                conn.close()

                ok = 0
                fail = 0
                for uid in user_ids:
                    try:
                        await context.bot.send_message(chat_id=uid, text=preview)
                        ok += 1
                        await asyncio.sleep(0.03)
                    except RetryAfter as e:
                        await asyncio.sleep(e.retry_after + 1)
                        try:
                            await context.bot.send_message(chat_id=uid, text=preview)
                            ok += 1
                        except Exception:
                            fail += 1
                    except Exception:
                        fail += 1

                await q.message.reply_text(
                    "✅ تم الإرسال.\n"
                    f"• تم الإرسال إلى: {ok}\n"
                    f"• فشل الإرسال إلى: {fail}"
                )
                await reply_plain(q.message, "👑 لوحة الإحصائيات (للأدمن فقط):", reply_markup=admin_menu())
                return
# regular bot callbacks
        if data.startswith("dept:"):
            idx = int(data.split(":")[1])
            if 0 <= idx < len(departments):
                name = departments[idx]
                num = phonebook.get(normalize_arabic(name), "")
                if update.effective_user:
                    upsert_user(update.effective_user)
                    log_event("dept_select", update.effective_user.id, chat_id, dept=name)
                await q.answer(text=f"{name}: {num if num else '—'}", show_alert=False)
                await reply_plain(q.message, f"📞 {name} — {num if num else '—'}")
            else:
                await q.answer("خيار غير صالح.", show_alert=False)
            return

        if data.startswith("allp:"):
            page = int(data.split(":")[1])
            await q.answer()
            await safe_edit(q, "اختر القسم من القائمة:", reply_markup=grid_all(page))
            return

        if data.startswith("srchp:"):
            page = int(data.split(":")[1])
            matches = context.user_data.get("last_search_indices", [])
            await q.answer()
            await safe_edit(q, "🔎 تم العثور على عدة نتائج، اختر القسم:", reply_markup=grid_search(matches, page))
            return

        if data == "home":
            await q.answer()
            try:
                await q.message.edit_text(build_intro(), reply_markup=None)
            except:
                pass
            await reply_plain(q.message, "رجعت إلى القائمة الرئيسية.", reply_markup=MAIN_KB)
            return

        if data == "noop":
            await q.answer()
            return

        await q.answer()

    except Exception:
        try:
            await q.answer("صار خطأ بسيط، جرّب مرة ثانية.", show_alert=False)
        except:
            pass

# ---------------- تشغيل ----------------
def read_token() -> Optional[str]:
    tok = os.getenv("TELEGRAM_BOT_TOKEN")
    if tok:
        return tok.strip()
    path = os.path.join(BASE, "token.txt")
    if os.path.exists(path):
        return open(path, "r", encoding="utf-8").read().strip()
    return None

if __name__ == "__main__":
    init_db()
    n, msg = load_phonebook()
    logging.info(msg)

    token = read_token()
    if not token:
        print("❌ لا يوجد توكن: ضع TELEGRAM_BOT_TOKEN أو token.txt.")
        raise SystemExit(1)

    app = ApplicationBuilder().token(token).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("about", about_cmd))
    app.add_handler(CommandHandler("reload", reload_cmd))
    app.add_handler(CommandHandler("admin", admin_cmd))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(CallbackQueryHandler(on_callback))

    print("📞 PhoneBook Bot running…")
    app.run_polling()
